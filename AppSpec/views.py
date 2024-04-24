from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.urls import reverse_lazy

from .models import *
from .forms import *

from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth import authenticate, login

import openpyxl
from openpyxl.styles import Alignment
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django.views.generic import ListView, CreateView, UpdateView, DeleteView

from datetime import datetime, timedelta
from django.contrib import messages
from django.shortcuts import render
from django.http import JsonResponse

# Create your views here.
# ------------------------------------------------------------------------HOME
def home(request):
    return render(request, "AppSpec/home.html")

# ------------------------------------------------------------------------SUMAR HORAS
def sumar_horas_y_minutos(hora1, hora2):
    horas1, minuto1 = map(int, hora1.split(':'))
    horas2, minuto2 = map(int, hora2.split(':'))
    minutos_total = minuto1 + minuto2
    horas_totales = horas1 + horas2 + (minutos_total // 60)
    minutos_total %= 60
    return f"{horas_totales}:{minutos_total}"

# ------------------------------------------------------------------------AREA
class AreaList(LoginRequiredMixin, ListView):
    model = Area
    template = "AppSpec/areas_list.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['areas_con_personas'] = []

        # Iterar sobre todas las áreas
        for area in context['object_list']:
            # Obtener todas las personas asociadas a esta área
            personas_en_area = area.personas.all()
            
            # Agregar el área y las personas a la lista de context
            context['areas_con_personas'].append({'area': area, 'personas': personas_en_area})

        return context


class AreaCreate(LoginRequiredMixin, CreateView):
    model = Area
    fields = ['nombre']
    success_url = reverse_lazy('areas')


class AreaUpdate(LoginRequiredMixin, UpdateView):
    model = Area
    fields = ['nombre']
    success_url = reverse_lazy('areas')


class AreaDelete(LoginRequiredMixin, DeleteView):
    model = Area
    success_url = reverse_lazy('areas')
    
# ------------------------------------------------------------------------PERSONA
class PersonaList(LoginRequiredMixin, ListView):
    model = Persona


class PersonaCreate(LoginRequiredMixin, CreateView):
    model = Persona
    fields = ['nombre', 'num_tarjeta', 'area']
    success_url = reverse_lazy('personas')


class PersonaUpdate(LoginRequiredMixin, UpdateView):
    model = Persona
    fields = ['area']
    success_url = reverse_lazy('personas')


class PersonaDelete(LoginRequiredMixin, DeleteView):
    model = Persona
    success_url = reverse_lazy('personas')


@login_required
def buscar_personas(request):
    if "buscar" in request.GET:
        patron = request.GET["buscar"]
        if patron:
            personas = Persona.objects.filter(nombre__icontains=patron)
            contexto = {"persona_list": personas}
            return render(request, "AppSpec/persona_list.html", contexto)
    return HttpResponse("No se ingresaron patrones de búsqueda")

# ------------------------------------------------------------------------LOGIN
def login_request(request):
    if request.method == "POST":
        usuario = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=usuario, password=password)
        if user is not None:
            login(request, user)
            return redirect("home")
        else:
            error_message = "Credenciales incorrectas. Por favor, inténtalo de nuevo."
            return render(request, "AppSpec/login.html", {"form": AuthenticationForm(), "error_message": error_message})
    else:
        
        return render(request, "AppSpec/login.html", {"form": AuthenticationForm()})
    
# ------------------------------------------------------------------------SELECCIONAR, IMPORTAR, CARGAR EXCEL
@login_required
def seleccionar_archivo(request):
    return render(request, 'AppSpec/seleccionarArchivo.html')


@login_required
def import_excel(request):
    if request.method == 'POST' and request.FILES['archivo_excel']:
        archivo_excel = request.FILES['archivo_excel']
        workbook = openpyxl.load_workbook(archivo_excel)
        dataframe = workbook.active
        filtered_data = {}

        for row in range(2, dataframe.max_row + 1):
            num_tarjeta = dataframe.cell(row=row, column=1).value
            nombre = dataframe.cell(row=row, column=2).value
            fecha_hora_cell = dataframe.cell(row=row, column=3)
            fecha_hora_value = fecha_hora_cell.value

            if fecha_hora_value is None:
                continue

            # Convertir la fecha y hora en formato "%Y%m%d%H%M%S" a un objeto datetime
            fecha_hora_str = str(fecha_hora_value)
            fecha_hora = datetime.strptime(fecha_hora_str, "%Y%m%d%H%M%S")

            # Actualizar el diccionario con los datos filtrados
            key = (num_tarjeta, fecha_hora.strftime("%Y%m%d"))  # Usamos solo la fecha como clave
            if key not in filtered_data:
                filtered_data[key] = {"nombre": nombre,"num_tarjeta": num_tarjeta, "primera_marcacion": fecha_hora, "ultima_marcacion": fecha_hora}
            else:
                if fecha_hora < filtered_data[key]["primera_marcacion"]:
                    filtered_data[key]["primera_marcacion"] = fecha_hora
                if fecha_hora > filtered_data[key]["ultima_marcacion"]:
                    filtered_data[key]["ultima_marcacion"] = fecha_hora

        for key, marcaciones in filtered_data.items():
            # Calcular la diferencia de tiempo entre la primera y la última marcación
            diferencia_tiempo = marcaciones["ultima_marcacion"] - marcaciones["primera_marcacion"]

            # Ajustar la diferencia de tiempo si cruza la medianoche
            if diferencia_tiempo.days < 0:
                diferencia_tiempo = timedelta(days=1) + diferencia_tiempo

            # Extraer la cantidad total de minutos de la diferencia de tiempo
            total_minutos = int(diferencia_tiempo.total_seconds() / 60)  # Convertir segundos a minutos

            # Redondear los minutos hacia abajo
            horas = total_minutos // 60
            minutos = total_minutos % 60

            # Formatear las horas y minutos en una cadena
            hora_completa = f"{horas:02d}:{minutos:02d}"

            # Agregar el total de horas y minutos al diccionario
            marcaciones["total_horas"] = hora_completa
            
        datos_con_personas = []
        mensaje = {}
        for key, marcaciones in filtered_data.items():
        # Obtener los datos de la marcación actual
            nombre = marcaciones["nombre"]
            num_tarjeta = marcaciones["num_tarjeta"]
            primera_marcacion = marcaciones["primera_marcacion"]
            ultima_marcacion = marcaciones["ultima_marcacion"]
            total_horas = marcaciones["total_horas"]
            

            # Buscar la persona correspondiente en la base de datos
            try:
                persona = Persona.objects.get(num_tarjeta=num_tarjeta)
                datos_con_personas.append({
                'area': persona.area.nombre,
                'nombre_persona': nombre,
                'num_tarjeta': num_tarjeta,
                'primera_marcacion': primera_marcacion,
                'ultima_marcacion': ultima_marcacion,
                'total_horas': total_horas
            })
            except Persona.DoesNotExist:
            # Manejar el caso en el que no se encuentre la persona en la base de datos
             """   mensaje.update({nombre: num_tarjeta})
                
        if len(mensaje) != 0:
            mensajito = ""
            for key, value in mensaje.items():
                mensajito += f'La persona {key} con número de tarjeta {value} no está registrada en la base de datos. Por favor, agregue la información correspondiente.<br>'
            return HttpResponse(mensajito)"""
        datos_con_personas_ordenado = sorted(datos_con_personas, key=lambda x: x['area'])
        #datos_agrupados_por_area = defaultdict(list)
        #for data in datos_con_personas_ordenado:
            #datos_agrupados_por_area[data['area']].append(data)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['Área', 'Nombre de la persona','Nº Tarjeta', 'Primera Marcacion', 'Ultima Marcacion', 'Horas'])
        prev_area = None
        total_de_horas = {}
        for area in datos_con_personas_ordenado:
            if area['area'] != prev_area:
                ws.append([area['area'], area['nombre_persona'],area['num_tarjeta'],area['primera_marcacion'],area['ultima_marcacion'], area['total_horas']])    
                prev_area = area['area']
                total_de_horas.update({area['nombre_persona'] : area['total_horas']})
            else:
                ws.append(['', area['nombre_persona'],area['num_tarjeta'],area['primera_marcacion'],area['ultima_marcacion'], area['total_horas']])
                if total_de_horas.get(area['nombre_persona']):
                    total_de_horas[area['nombre_persona']] = sumar_horas_y_minutos(total_de_horas[area['nombre_persona']], area['total_horas'])
                else:
                    total_de_horas.update({area['nombre_persona'] : area['total_horas']})
        new_hoja = wb.create_sheet("Total de horas")
        for idx, (clave, valor) in enumerate(total_de_horas.items(), start=1):
            new_hoja[f"A{idx}"] = clave
            new_hoja[f"B{idx}"] = valor
        column_widths = [60, 300, 70, 150, 150,50]  # Ancho de las columnas en píxeles
        for i, width in enumerate(column_widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width / 7

        for row in ws.iter_rows(min_row=1):  # Comenzar desde la segunda fila para evitar la cabecera
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=datos.xlsx'
        
        wb.save(response)

        return response
        

    # Renderiza la plantilla con los datos procesados
        #return render(request, 'AppSpec/resultado_importacion.html', {'datos_con_personas': datos_con_personas_ordenado})

    return render(request, 'AppSpec/import_excel.html')


# @login_required
# def cargar_excel(request):
#     if request.method == 'POST' and request.FILES.get('archivo_excel'):
#         archivo_excel = request.FILES['archivo_excel']
        
#         try:
#             # Verificar si el archivo es un archivo Excel
#             if archivo_excel.name.endswith('.xlsx'):
#                 # Abrir el archivo Excel
#                 workbook = openpyxl.load_workbook(archivo_excel)
#                 worksheet = workbook.active
                

#                 # Iterar sobre las filas del archivo Excel
#                 for row in worksheet.iter_rows(min_row=2, values_only=True):
#                     # Obtener el nombre del área de la fila actual
#                     nombre_area = row[2]  # Suponiendo que el nombre del área está en la tercera columna

#                     # Buscar la instancia de Area correspondiente al nombre leído del archivo Excel
#                     try:
#                         area = Area.objects.get(nombre=nombre_area)
#                     except Area.DoesNotExist:
#                         # Si no se encuentra el área, crear una nueva instancia de Area
#                         area = Area(nombre=nombre_area)
#                         area.save()

#                     # Crear un objeto del modelo Persona y asignar los valores de las columnas
#                     objeto = Persona(
#                         nombre=row[0],
#                         num_tarjeta=row[1],
#                         area=area,
#                         # Añade más campos según sea necesario
#                     )
#                     objeto.save()

#                 messages.success(request, 'Los datos se han cargado correctamente desde el archivo Excel.')
#             else:
#                 messages.error(request, 'El archivo seleccionado no es un archivo Excel válido.')
#         except Exception as e:
#             messages.error(request, f'Error al cargar datos desde el archivo Excel: {e}')

#     return render(request, 'AppSpec/cargar_excel.html')

# ------------------------------------------------------------------------CARGAR PERSONAS
@login_required
def obtener_personas_area(request):
    if request.method == 'GET' and 'area_id' in request.GET:
        area_id = request.GET['area_id']
        try:
            area = Area.objects.get(id=area_id)
            personas = area.personas.all()
            personas_data = [{'nombre': persona.nombre, 'num_tarjeta': persona.num_tarjeta} for persona in personas]
            return JsonResponse({'personas': personas_data})
        except Area.DoesNotExist:
            return JsonResponse({'error': 'El área seleccionada no existe'}, status=404)
    return JsonResponse({'error': 'Se requiere el ID del área'}, status=400)
